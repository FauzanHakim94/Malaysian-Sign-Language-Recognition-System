from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import Workbook, load_workbook
import cv2
import os
import time

import ui.Message as Message
import utils.ImageProcessing as IP

class RTab:
    def init(self, rootPath, alphabets, datasetPath):
        RTab.rootPath = rootPath
        RTab.alphabets = alphabets
        RTab.datasetPathRecord = datasetPath
        RTab.headerFields = ['File Name', 'Classes', 'Width', 'Height', 'X Min', 'X Max', 'Y Min', 'Y Max']
        RTab.handSide = ['R','L']      
        RTab.recordTab = self.tabs.findChild(QWidget, "RECORDTab")
        RTab.msg_instance = Message.Msg()

        # Message Box
        RTab.msg = QMessageBox()
        RTab.searchLineEdit = self.findChild(QLineEdit, "filePathEditLine")
        # Widget List
        RTab.list1 = RTab.recordTab.findChild(QListWidget, "listWidget1")
        RTab.scrollbar1 = RTab.recordTab.findChild(QScrollBar, "vScrollBar1")
        RTab.cameraButton = RTab.recordTab.findChild(QPushButton, "cameraButton")
        RTab.spinCameraID = RTab.recordTab.findChild(QSpinBox, "spinBoxCamera")
        RTab.snapButton = RTab.recordTab.findChild(QPushButton, "snapButton")
        RTab.liveImage = RTab.recordTab.findChild(QLabel, "liveImage")
        RTab.labelCaptureProgress = RTab.recordTab.findChild(QLabel, "labelCaptureProgress")
        RTab.capturedProgress = RTab.recordTab.findChild(QProgressBar, "capturedProgress")
            
        # Widget Camera Connection
        RTab.cameraID = 0
        RTab.cameraButtonState = 0
        RTab.snapButtonState = 0
        RTab.frameCounter = 0
        RTab.displayModelX = 0
        RTab.handCount = 0
        RTab.spinCameraID.valueChanged.connect(RTab.cameraIDChanged)        
        RTab.cameraButton.clicked.connect(RTab.cameraButtonClicker)
        RTab.snapButton.clicked.connect(RTab.snapButtonClicker)
        # Widget List Connection
        RTab.addList(self) 
        RTab.list1.itemClicked.connect(RTab.itemClickedEvent)
        # Display Initial Tab Progress
        RTab.labelCaptureProgress.setStyleSheet("background-color: white")
        RTab.labelCaptureProgress.clear()
        RTab.capturedProgress.setValue(0)

        # Initial Activity
        RTab.Worker1 = Worker1()
        RTab.Worker1.ImageUpdate.connect(RTab.ImageUpdateSlot)        

    """message box settings"""
    def callMessageBox(self, title, message):
        RTab.msg.setWindowTitle(title)
        RTab.msg.setText(message)
        RTab.msg.setIcon(QMessageBox.Critical)
        RTab.msg.setStandardButtons(QMessageBox.Close)
        RTab.msg.exec_()

    """when changing camera channel"""
    def cameraIDChanged(self):
        RTab.liveImage.setText(RTab.msg_instance.get_message(1))
        RTab.cameraID = RTab.spinCameraID.value()
        RTab.cameraButtonState = 0
        RTab.threadState(self)

    """when on/off camera button clicked"""
    def cameraButtonClicker(self):
        RTab.cameraButtonState = not RTab.cameraButtonState
        RTab.threadState(self)

    """when threadState is updated"""
    def threadState(self):
        if RTab.cameraButtonState:
            if not RTab.Worker1.isRunning():
                RTab.liveImage.clear()
                RTab.liveImage.setText(RTab.msg_instance.get_message(2))
                RTab.Worker1.start()
        else:
            if RTab.Worker1.isRunning():
                RTab.Worker1.stop()
                RTab.liveImage.clear()
                RTab.liveImage.setText(RTab.msg_instance.get_message(1))

    """when snap camera button clicked"""
    def snapButtonClicker(self):        
        # Camera is not turned on
        if not RTab.cameraButtonState:
            RTab.callMessageBox(self, RTab.msg_instance.get_message(1), RTab.msg_instance.get_message(3))
            return
        # Get selected gesture        
        index = RTab.datasetPathRecord.find("\Gesture")
        if index < 0:
            RTab.callMessageBox(self, RTab.msg_instance.get_message(5), RTab.msg_instance.get_message(6))
            return
        else:
            RTab.record_label(self, index)
        # Initial setup for thread
        RTab.snapButtonState = 1
        RTab.frameCounter = 0
        RTab.displayModelX = 1
        RTab.handCount = 0
        # Initial setup for progress bar
        RTab.capturedProgress.setRange(0, 5)
        RTab.capturedProgress.setValue(0)
        # Initial time for model image
        RTab.startTimeModel = time.time()

    """update the label of dataset"""
    def record_label(self, index):
        # Finding dataset file for recording
        tempDir = RTab.datasetPathRecord[:index] + "\Record Image Label.xlsx"        
        # Load dataset label excel
        if os.path.exists(tempDir):
            wb = load_workbook(tempDir)
        else:
            workbook = Workbook()
            sheet = workbook.active
            workbook.save(filename=tempDir)
            wb = load_workbook(tempDir)
        sheet = wb.worksheets[0]
        # Naming column title in spreadsheet
        for col in range(len(RTab.headerFields)):
            sheet.cell(row=1, column=col + 1).value = RTab.headerFields[col]
        wb.save(tempDir)

    """adding gesture type in the list"""
    def addList(self):
        for x in range(len(RTab.alphabets)):
            RTab.list1.addItem(QListWidgetItem("Gesture "+RTab.alphabets[x]))
        RTab.list1.setVerticalScrollBar(RTab.scrollbar1)

    """when gesture type is selected"""
    def itemClickedEvent(self):
        RTab.listRow = RTab.list1.currentRow()
        index = RTab.datasetPathRecord.find("dataset")
        RTab.datasetPathRecord = RTab.datasetPathRecord[:index+7]+'\Gesture '+RTab.alphabets[RTab.listRow]
        RTab.searchLineEdit.setText(RTab.datasetPathRecord)

    """update image for webcam"""
    def ImageUpdateSlot(image):
        RTab.liveImage.clear()
        RTab.liveImage.setPixmap(QPixmap.fromImage(image))

    """update text on webcam image"""
    def ImageUpdate2Slot(string):
        RTab.liveImage.clear()
        RTab.liveImage.setText(string)

"""Thread 1"""
class Worker1(QThread):
    # Slot to update
    ImageUpdate = pyqtSignal(QImage)

    """when thread is running"""
    def run(self):
        # Initialize camera
        if not IP.IProcessing.init(self, RTab.cameraID):
            RTab.cameraButtonState = 0
            self.ImageUpdate2.emit(RTab.msg_instance.get_message(4))
            self.ThreadActive = False
        else:
            self.ThreadActive = True
            model_directory = None
            frameCounter = 0

        # While camera is detected and open
        while self.ThreadActive:
            # Record frame from video
            ret, frame, Image, h, w, c = IP.IProcessing.read_camera(self)
            # Get model to display
            if RTab.displayModelX:
                model_directory, imgNo = self.display_model()
                if time.time() - RTab.startTimeModel > 5:
                        RTab.displayModelX = 0
                        RTab.startTime = time.time()
                else:
                    seconds = str(5-round(time.time() - RTab.startTimeModel))
            else:
                handResult = 0
                Image = IP.IProcessing.mediapipe_process(self, RTab.handCount)
            # Displaying image
            if ret:
                if RTab.displayModelX:
                    if model_directory is not None:
                        cv2.putText(img=model_directory, text=seconds, org=(1155, 708), fontFace=cv2.FONT_HERSHEY_TRIPLEX, fontScale=3, color=(0, 0, 255), thickness=3)
                        Pic = self.convert_cv_qt(model_directory)                
                else:
                    FlippedImage = cv2.flip(Image, 1)
                    ConvertToQtFormat = QImage(FlippedImage.data, FlippedImage.shape[1], FlippedImage.shape[0], QImage.Format_RGB888)                
                    Pic = ConvertToQtFormat.scaled(640, 480, Qt.KeepAspectRatio)
                if self.ThreadActive:
                    self.ImageUpdate.emit(Pic)
         
    """stop the thread"""
    def stop(self):
        self.ThreadActive = False
        self.Capture.release()
        self.terminate()

    """get model image"""
    def display_model(self):
        if not os.path.exists(RTab.datasetPathRecord):
            os.makedirs(RTab.datasetPathRecord)        
        if os.getcwd != RTab.datasetPathRecord:
            os.chdir(RTab.datasetPathRecord)
        imgNo = len(os.listdir())
        if RTab.listRow<24:
            model_directory = cv2.imread(RTab.rootPath +'/ui_assets' + '/Gesture ' + RTab.alphabets[RTab.listRow] + ' Model ' + RTab.handSide[RTab.handCount] + '.PNG')
        else:
            model_directory = cv2.imread(RTab.rootPath +'/ui_assets' + '/Gesture ' + RTab.alphabets[RTab.listRow]+'.PNG')
        return model_directory, imgNo
    
    """convert cv format to qt"""
    def convert_cv_qt(self, cv_img):
        # Convert from an opencv image to QPixmap
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(640, 480, Qt.KeepAspectRatio)
        return p  