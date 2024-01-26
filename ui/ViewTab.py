import os
import sys
import cv2
import time
import gc
from openpyxl import Workbook, load_workbook

from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.uic import loadUi
import mediapipe as mp

rootPathRecord = "null"
datasetPathRecord = "null"
V_alphabets = "null"
globalPath = "null"
row = 0

mpHolistic = 0
mpDraws = 0
holistic = 0

class VTab:
    def init(self, rootPath, datasetPath, alphabets):
        global rootPathRecord, datasetPathRecord, V_alphabets, mpHolistic, mpDraws, holistic

        rootPathRecord = rootPath
        datasetPathRecord = datasetPath
        V_alphabets = alphabets

        VTab.msg = QMessageBox()
        VTab.viewTab = self.tabs.findChild(QWidget, "VIEWTab")
        VTab.tree = VTab.viewTab.findChild(QTreeView, "treeView")
        VTab.nextButton = VTab.viewTab.findChild(QPushButton, "nextButton")
        VTab.prevButton = VTab.viewTab.findChild(QPushButton, "prevButton")
        VTab.recordedImage = VTab.viewTab.findChild(QLabel, "recordedImage")
        VTab.RoI_Image = VTab.viewTab.findChild(QLabel, "RoI_Image")
        VTab.input_FileName = VTab.viewTab.findChild(QLabel, "input_FileName")
        VTab.input_Class = VTab.viewTab.findChild(QLabel, "input_Class")
        VTab.input_Width = VTab.viewTab.findChild(QLabel, "input_Width")
        VTab.input_Height = VTab.viewTab.findChild(QLabel, "input_Height")
        VTab.input_Xmin = VTab.viewTab.findChild(QLabel, "input_Xmin")
        VTab.input_Xmax = VTab.viewTab.findChild(QLabel, "input_Xmax")
        VTab.input_Ymin = VTab.viewTab.findChild(QLabel, "input_Ymin")
        VTab.input_Ymax = VTab.viewTab.findChild(QLabel, "input_Ymax")

        VTab.initTreeModel(self, datasetPathRecord)

        #Widget Connection
        VTab.tree.clicked.connect(VTab.itemTreeSelected)
        VTab.nextButton.clicked.connect(VTab.nextButtonClicker)
        VTab.prevButton.clicked.connect(VTab.prevButtonClicker)

    def callMessageBox(self, title, message):
        VTab.msg.setWindowTitle(title)
        VTab.msg.setText(message)
        VTab.msg.setIcon(QMessageBox.Critical)
        VTab.msg.setStandardButtons(QMessageBox.Close)
        VTab.msg.exec_()

    def initTreeModel(self, datasetPathRecord):
        VTab.model = QFileSystemModel()
        VTab.model.setRootPath(datasetPathRecord)
        VTab.tree.setModel(VTab.model)
        VTab.tree.setRootIndex(VTab.model.index(datasetPathRecord))
        VTab.tree.setAnimated(False)
        VTab.tree.setIndentation(20)
        VTab.tree.setSortingEnabled(False)
        VTab.tree.setWindowTitle("Dir View")
        VTab.tree.resize(640, 480)

    def nextButtonClicker(self):
        global globalPath
        fIndex = VTab.tree.currentIndex()
        fIndex = VTab.tree.indexBelow(fIndex)
        path = VTab.model.filePath(fIndex)
        imgLog = path.find('.jpg')
        if imgLog>=0:
            VTab.updateRecordedImage(self, path)
        VTab.tree.setCurrentIndex(fIndex)
        globalPath = path

    def prevButtonClicker(self):
        global globalPath
        fIndex = VTab.tree.currentIndex()
        fIndex = VTab.tree.indexAbove(fIndex)
        path = VTab.model.filePath(fIndex)
        imgLog = path.find('.jpg')
        if imgLog>=0:
            VTab.updateRecordedImage(self, path)
        VTab.tree.setCurrentIndex(fIndex)
        globalPath = path

    def itemTreeSelected(self):
        global globalPath
        fIndex = VTab.tree.currentIndex()
        path = VTab.model.filePath(fIndex)
        imgLog = path.find('.jpg')
        if imgLog >= 0:
            VTab.updateRecordedImage(self, path)
            globalPath = path

    def updateRecordedImage(self, path):
        global row, holistic, rootPathRecord
        img = cv2.imread(path)
        h, w, c = img.shape
        excelName = '\Record Image Label.xlsx'
        print("Test0")
        if VTab.callWorkbook(self, excelName):
            index = path.find('/Gesture')
            print("Test0.1")
            imageName = path[index+9:]
            print("Test0.1")
            index = imageName.find('/')
            print("Test0.1")
            row, dataImageName, dataImageClass, dataImageWidth, dataImageHeight, dataXY = VTab.readWorkbook(self, excelName, imageName[index+1:])
            print("Test0.1")
            print(path+str(w)+str(h)+str(dataXY[0])+str(dataXY[1])+str(dataXY[2])+str(dataXY[3]))
            VTab.updateSummary(self, path, w, h, dataXY[0], dataXY[1], dataXY[2], dataXY[3])
            print("Test0.2")
            VTab.recordedImage.clear()
            print("Test0.3")
            rec_img = VTab.convert_cv_qt(self, 0, img, 100, 100)
            print("Test0.4")
            VTab.recordedImage.setPixmap(QPixmap.fromImage(rec_img))
            print("Test1")
            index = rootPathRecord.find("66. Malaysia Sign Language Recognition System")
            tempDir = rootPathRecord[:index + 45] + "\landmarkDataset"+"\\"
            print(imageName)
            if imageName[1:2] == '/':
                roi_img = cv2.imread(tempDir+imageName[2:])
            else:
                Nindex = imageName.find('/')
                roi_img = cv2.imread(tempDir+imageName[Nindex:])
            VTab.RoI_Image.clear()
            roi_img = cv2.flip(roi_img,1)
            roi_img = VTab.convert_cv_qt(self, 0, roi_img, 100, 100)
            VTab.RoI_Image.setPixmap(QPixmap.fromImage(roi_img))
            print("Test2")


    def updateSummary(self, path, wd, ht, xmin, xmax, ymin, ymax):
        index = path.find("dataset")
        path = path[index + 8:]
        index = path.find("/")
        fn = path[index+1:]
        cl = path[:index]

        VTab.input_FileName.setText(fn)
        VTab.input_Class.setText(cl)
        VTab.input_Width.setText(str(wd))
        VTab.input_Height.setText(str(ht))
        VTab.input_Xmin.setText(str(xmin))
        VTab.input_Xmax.setText(str(xmax))
        VTab.input_Ymin.setText(str(ymin))
        VTab.input_Ymax.setText(str(ymax))

    def callWorkbook(self, excelName):
        global datasetPathRecord

        tempDir = datasetPathRecord + excelName
        if not os.path.exists(tempDir):
            VTab.callMessageBox(self, "The excel file is not found", "Please record images first")
            return False
        return True

    def readWorkbook(self, excelName, imageName):
        global datasetPathRecord

        tempDir = datasetPathRecord + excelName
        wb = load_workbook(tempDir)
        sheet = wb.worksheets[0]
        lastRow = sheet.max_row
        dataXY = {}
        for elements in range(lastRow):
            exImageName = sheet.cell(row=elements+1,column=1).value
            if exImageName == imageName:
                row = elements+1
                dataImageName = exImageName
                dataImageClass = sheet.cell(row=elements+1,column=1).value
                dataImageWidth = sheet.cell(row=elements+1,column=2).value
                dataImageHeight = sheet.cell(row=elements+1,column=3).value

                dataXY[0] = sheet.cell(row=elements+1,column=5).value
                dataXY[1] = sheet.cell(row=elements+1,column=6).value
                dataXY[2] = sheet.cell(row=elements+1,column=7).value
                dataXY[3] = sheet.cell(row=elements+1,column=8).value
                print(tempDir)
                wb.save(tempDir)
                print(dataXY[0], dataXY[1], dataXY[2], dataXY[3])
                return row,dataImageName,dataImageClass,dataImageWidth,dataImageHeight,dataXY

    def editWorkbook(self, excelName, rowX, dataImageXmin, dataImageXmax, dataImageYmin, dataImageYmax):
        global datasetPathRecord

        tempDir = datasetPathRecord + excelName
        wb = load_workbook(tempDir)
        sheet = wb.worksheets[0]
        sheet.cell(row=rowX, column=5).value = dataImageXmin
        sheet.cell(row=rowX, column=6).value = dataImageXmax
        sheet.cell(row=rowX, column=7).value = dataImageYmin
        sheet.cell(row=rowX, column=8).value = dataImageYmax
        wb.save(tempDir)

    def convert_cv_qt(self, flag, rgb_image, scaledWidth, scaledHeight):
        """Convert from an opencv image to QPixmap"""
        rgb_image = cv2.cvtColor(rgb_image, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        #p = convert_to_Qt_format.scaled(scaledWidth, scaledHeight, Qt.KeepAspectRatio)
        return convert_to_Qt_format